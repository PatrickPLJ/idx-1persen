#!/usr/bin/env python3
"""
Loop "Cek 1% Kepemilikan Saham IDX" — mesin fetch + diff + hitung fakta.

Sumber: https://www.idx.co.id/id/perusahaan-tercatat/data-kepemilikan-saham/
File target: peng-XX-XXXXX-satu-persen.xlsx (terbit bulanan, berisi semua
investor dengan kepemilikan >= 1% per emiten).

Perintah:
  python3 fetch_idx.py check [--dry-run]   # cek data baru, hitung fakta
  python3 fetch_idx.py mark <id> <delivered|rejected> [alasan]
  python3 fetch_idx.py status              # lihat isi state

Guardrail (di level script):
  - Maksimal 1 output per hari (cek last_output_date di state).
  - Script ini TIDAK mengirim apa pun keluar, TIDAK menghapus data,
    TIDAK mengeluarkan uang. Hanya baca web IDX + tulis file lokal.
"""
import json
import re
import sys
import time
import datetime
from pathlib import Path

import cloudscraper
import openpyxl

BASE = Path(__file__).resolve().parent
STATE_FILE = BASE / "1% kepemilikan saham.json"
DATA_DIR = BASE / "data"
FACTS_DIR = BASE / "facts"
PAGE_URL = "https://www.idx.co.id/id/perusahaan-tercatat/data-kepemilikan-saham/"
TOP_N = 15  # batas daftar top movers di facts


def now_iso():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today():
    return datetime.date.today().isoformat()


def load_state():
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text())
    return {
        "keterangan": "State file loop 'Cek 1% Kepemilikan Saham IDX' — "
                      "daftar pengumuman yang sudah diproses. Jangan edit manual.",
        "dibuat": now_iso(),
        "processed": {},
        "last_run": None,
        "last_output_date": None,
    }


def save_state(state):
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False))


def ambil_url(url, percobaan=4, timeout=90):
    """GET dengan retry — Cloudflare IDX kadang menolak (403) secara acak."""
    terakhir = None
    for i in range(percobaan):
        try:
            s = cloudscraper.create_scraper()
            r = s.get(url, timeout=timeout)
            r.raise_for_status()
            return r
        except Exception as e:
            terakhir = e
            if i < percobaan - 1:
                time.sleep(8 * (i + 1))
    raise terakhir


def scrape_satu_persen_list():
    """Ambil daftar file satu-persen dari halaman IDX, urutan terbaru dulu."""
    r = ambil_url(PAGE_URL)
    html = r.text.replace("\\u002F", "/")
    urls = re.findall(
        r'https://www\.idx\.co\.id/Media/([A-Za-z0-9]+)/(peng-[\w\-]*satu-persen)\.xlsx',
        html,
    )
    items, seen = [], set()
    for slug, stem in urls:
        item_id = f"{stem}@{slug}"
        if item_id in seen:
            continue
        seen.add(item_id)
        items.append({
            "id": item_id,
            "filename": f"{stem}.xlsx",
            "url": f"https://www.idx.co.id/Media/{slug}/{stem}.xlsx",
        })
    if not items:
        raise RuntimeError("Tidak menemukan file satu-persen di halaman IDX — "
                           "struktur halaman mungkin berubah.")
    return items  # urutan halaman = terbaru dulu


def download(item, scraper=None):
    dest = DATA_DIR / item["filename"]
    if dest.exists() and dest.stat().st_size > 0:
        return dest
    r = ambil_url(item["url"], timeout=120)
    if not r.content[:2] == b"PK":
        raise RuntimeError(f"Unduhan {item['url']} bukan file xlsx valid.")
    dest.write_bytes(r.content)
    return dest


STOPWORDS = {"PT", "TBK", "PERSERO", "PERUSAHAAN", "PERSEROAN"}


def normalize_investor(name):
    """Samakan variasi penulisan nama KSEI antar bulan.

    'PT. BUMI SERPONG DAMAI Tbk' dan 'BUMI SERPONG DAMAI TBK PT' adalah
    entitas yang sama — tanpa ini, diff bulanan penuh false positive.
    """
    n = re.sub(r"[.,()'\"]", " ", name.upper())
    tokens = [t for t in n.split() if t not in STOPWORDS]
    return " ".join(tokens)


def parse_xlsx(path):
    """Baca xlsx satu-persen -> (tanggal_data, dict {(kode, investor_norm): row}).

    Baris dengan kode emiten + nama ternormalisasi sama dijumlahkan
    (KSEI kadang memecah satu investor jadi beberapa baris).
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_seen = False
    holdings = {}
    tanggal = None
    for row in ws.iter_rows(values_only=True):
        if not header_seen:
            if row[0] == "DATE":
                header_seen = True
            continue
        if row[0] is None or row[1] is None:
            continue
        tanggal = str(row[0])[:10]
        kode = str(row[1]).strip()
        investor = str(row[3]).strip() if row[3] else "?"
        try:
            pct = float(row[11]) if row[11] is not None else None
        except (TypeError, ValueError):
            pct = None
        try:
            shares = int(row[10]) if row[10] is not None else None
        except (TypeError, ValueError):
            shares = None
        key = (kode, normalize_investor(investor))
        if key in holdings:
            h = holdings[key]
            if pct is not None:
                h["pct"] = round((h["pct"] or 0) + pct, 4)
            if shares is not None:
                h["shares"] = (h["shares"] or 0) + shares
        else:
            holdings[key] = {
                "display": investor,
                "issuer": str(row[2]).strip() if row[2] else "",
                "klasifikasi": str(row[4]).strip() if row[4] else "",
                "lokal_asing": str(row[5]).strip() if row[5] else "",
                "shares": shares,
                "pct": pct,
            }
    wb.close()
    if not holdings:
        raise RuntimeError(f"File {path.name} tidak berisi data yang bisa dibaca.")
    return tanggal, holdings


def compute_facts(cur_id, cur_path, prev_id, prev_path):
    cur_date, cur = parse_xlsx(cur_path)
    facts = {
        "dihitung": now_iso(),
        "sumber": {"sekarang": cur_id, "pembanding": prev_id},
        "tanggal_data": cur_date,
        "ringkasan": {
            "total_baris": len(cur),
            "jumlah_emiten": len({k for k, _ in cur}),
            "investor_lokal": sum(1 for v in cur.values() if v["lokal_asing"] == "L"),
            "investor_asing": sum(1 for v in cur.values() if v["lokal_asing"] == "F"),
        },
    }
    if prev_path is None:
        facts["pembanding"] = None
        return facts

    prev_date, prev = parse_xlsx(prev_path)
    cur_keys, prev_keys = set(cur), set(prev)
    masuk = sorted(cur_keys - prev_keys, key=lambda k: -(cur[k]["pct"] or 0))
    keluar = sorted(prev_keys - cur_keys, key=lambda k: -(prev[k]["pct"] or 0))

    # Pasangan keluar+masuk di emiten sama dengan pct nyaris identik
    # hampir pasti ganti nama pencatatan/kustodian, bukan transaksi nyata.
    renames, keluar_sisa, masuk_bersih = [], list(keluar), []
    for k in masuk:
        cand = [k2 for k2 in keluar_sisa if k2[0] == k[0]
                and cur[k]["pct"] is not None and prev[k2]["pct"] is not None
                and abs(cur[k]["pct"] - prev[k2]["pct"]) < 0.1]
        if cand:
            k2 = min(cand, key=lambda x: abs(cur[k]["pct"] - prev[x]["pct"]))
            keluar_sisa.remove(k2)
            renames.append({"kode": k[0], "dari": prev[k2]["display"],
                            "menjadi": cur[k]["display"], "pct": cur[k]["pct"]})
        else:
            masuk_bersih.append(k)
    masuk, keluar = masuk_bersih, keluar_sisa
    sama = [k for k in cur_keys & prev_keys
            if cur[k]["pct"] is not None and prev[k]["pct"] is not None
            and abs(cur[k]["pct"] - prev[k]["pct"]) >= 0.05]
    delta = sorted(sama, key=lambda k: cur[k]["pct"] - prev[k]["pct"])

    def fmt(keys, src, with_delta=False):
        out = []
        for k in keys[:TOP_N]:
            d = {
                "kode": k[0], "emiten": src[k]["issuer"],
                "investor": src[k]["display"],
                "klasifikasi": src[k]["klasifikasi"],
                "lokal_asing": src[k]["lokal_asing"],
                "pct": src[k]["pct"], "shares": src[k]["shares"],
            }
            if with_delta:
                d["pct_sebelum"] = prev[k]["pct"]
                d["pct_sesudah"] = cur[k]["pct"]
                d["delta_pct"] = round(cur[k]["pct"] - prev[k]["pct"], 2)
            out.append(d)
        return out

    cur_emiten = {k for k, _ in cur_keys}
    prev_emiten = {k for k, _ in prev_keys}
    facts["pembanding"] = {
        "tanggal_data_sebelum": prev_date,
        "total_baris_sebelum": len(prev),
        "jumlah_emiten_sebelum": len(prev_emiten),
        "jumlah_investor_masuk": len(masuk),
        "jumlah_investor_keluar": len(keluar),
        "jumlah_kemungkinan_ganti_nama": len(renames),
        "contoh_ganti_nama": sorted(renames, key=lambda r: -(r["pct"] or 0))[:10],
        "emiten_baru": sorted(cur_emiten - prev_emiten),
        "emiten_hilang": sorted(prev_emiten - cur_emiten),
        "top_investor_masuk": fmt(masuk, cur),
        "top_investor_keluar": fmt(keluar, prev),
        "top_kenaikan": fmt(list(reversed(delta)), cur, with_delta=True),
        "top_penurunan": fmt(delta, cur, with_delta=True),
    }
    return facts


def cmd_check(dry_run):
    state = load_state()
    state["last_run"] = now_iso()

    # Guardrail: maksimal 1 output per hari
    if not dry_run and state.get("last_output_date") == today():
        print("STATUS: DAILY_LIMIT")
        print("INFO: sudah ada output hari ini, guardrail max 1 output/hari aktif.")
        save_state(state)
        return

    items = scrape_satu_persen_list()
    print(f"INFO: {len(items)} file satu-persen ditemukan di halaman IDX.")

    # Seeding run pertama: catat semua KECUALI yang terbaru sebagai 'seeded'
    # supaya backlog lama tidak diproses satu-satu.
    if not state["processed"]:
        for it in items[1:]:
            state["processed"][it["id"]] = {
                "status": "seeded", "tanggal": now_iso(),
                "alasan": "run pertama — arsip lama dicatat tanpa diproses",
            }
        print(f"INFO: run pertama — {len(items) - 1} arsip lama ditandai 'seeded'.")

    baru = [it for it in items if it["id"] not in state["processed"]]
    if not baru:
        print("STATUS: NO_NEW_DATA")
        if not dry_run:
            save_state(state)
        return

    target = baru[-1]  # proses yang paling lama dulu (urut kronologis)
    idx = next(i for i, it in enumerate(items) if it["id"] == target["id"])
    pembanding = items[idx + 1] if idx + 1 < len(items) else None

    s = cloudscraper.create_scraper()
    cur_path = download(target, s)
    prev_path = download(pembanding, s) if pembanding else None
    print(f"INFO: diunduh {cur_path.name}"
          + (f" + pembanding {prev_path.name}" if prev_path else " (tanpa pembanding)"))

    facts = compute_facts(target["id"], cur_path,
                          pembanding["id"] if pembanding else None, prev_path)
    bulan = re.search(r"peng-(\d+)", target["filename"]).group(1)
    tahun = facts["tanggal_data"][:4]
    facts_path = FACTS_DIR / f"facts-{tahun}-{bulan}.json"
    facts_path.write_text(json.dumps(facts, indent=2, ensure_ascii=False))

    if not dry_run:
        save_state(state)

    print("STATUS: NEW_DATA")
    print(f"ID: {target['id']}")
    print(f"XLSX: {cur_path}")
    print(f"FACTS: {facts_path}")
    print(f"PERIODE: {facts['tanggal_data']}")
    if dry_run:
        print("INFO: dry-run — state TIDAK diubah.")


def cmd_mark(item_id, status, alasan):
    if status not in ("delivered", "rejected"):
        sys.exit("Status harus 'delivered' atau 'rejected'.")
    state = load_state()
    state["processed"][item_id] = {
        "status": status, "tanggal": now_iso(), "alasan": alasan or None,
    }
    if status == "delivered":
        state["last_output_date"] = today()
    save_state(state)
    print(f"OK: {item_id} -> {status}")


def cmd_status():
    state = load_state()
    print(json.dumps(state, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] == "check":
        cmd_check(dry_run="--dry-run" in args)
    elif args[0] == "mark" and len(args) >= 3:
        cmd_mark(args[1], args[2], " ".join(args[3:]))
    elif args[0] == "status":
        cmd_status()
    else:
        sys.exit(__doc__)
